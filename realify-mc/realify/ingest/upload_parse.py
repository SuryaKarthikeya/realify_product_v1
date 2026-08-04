"""Tolerant multi-store catalog upload parser (Build 1).

Accepts real store exports as CSV or XLSX, with messy/scraped headers
(e.g. 'Title__title__z5HRm', 'Price__whole__mQGs5 2'), and maps them to the
minimal seed shape the pipeline already understands:
    {asin, title, price, category, cogs, channel}

Design rules:
  - Column matching is fuzzy and order-free (substring + alias table).
  - Only ASIN is hard-required. price/title/category default gracefully.
  - COGS is OPTIONAL: if absent we derive it from price (DEFAULT_COGS_RATIO)
    so store exports without cost data still load instead of being silently
    dropped (the catalog-wipe failure mode we explicitly guard against).
  - Multiple files -> one logical catalog, each row tagged with its channel/store.
"""
import csv, io, re

DEFAULT_COGS_RATIO = 0.42          # cogs assumed when an export omits cost
DEFAULT_CATEGORY   = "Other Accessories"

# alias -> canonical field. Matched as case-insensitive substrings against the
# normalized header, declared order = priority (most specific / most correct first) —
# same convention as report_parse.classify()'s "ordered: most specific first" if-chain.
# 'fnsku' sits last: it's an Amazon-internal fulfillment ID, never what a tenant's
# known SKU/ASIN is keyed by, so it must lose to 'asin'/'sku' whenever both are present.
_ALIASES = {
    "asin":     ["asin", "product id", "productid", "sku id", "sku", "item id", "fnsku"],
    "title":    ["title", "product name", "item name", "name", "product", "description"],
    "price":    ["selling price", "sale price", "list price", "price whole", "price", "mrp", "rate", "amount"],
    "cogs":     ["cost of goods", "unit cost", "landed cost", "cogs", "cost price", "cost"],
    "category": ["amazon category", "category", "product type", "ptype", "type", "dept", "department"],
    "channel":  ["channel", "marketplace", "store", "platform", "source"],
}

def _norm(h):
    h = (h or "").strip().lower()
    h = re.sub(r"__[a-z0-9 ]+$", "", h)          # strip scraped suffix like __title__z5hrm
    h = re.sub(r"[^a-z0-9 ]+", " ", h)
    return re.sub(r"\s+", " ", h).strip()

def _find_header_row(matrix, aliases, window=20, min_hits=2):
    """First row (within the first `window` rows) with at least `min_hits` DISTINCT cells
    matching an alias substring; else 0. Real exports commonly carry several lines of
    banner/legend text before the real header row (e.g. Amazon's Unified Transaction Report
    has ~12) — without this, row 1 gets treated as the header and every column-match
    downstream silently fails on garbage text. Requiring >=2 cell-hits (not just any single
    substring hit anywhere in the row) matters: a banner line is one prose cell that can
    easily contain one incidental alias substring (e.g. a legend line defining 'Date/Time'
    contains the bare word 'date'), but a real header row has many distinct column names each
    matching some alias — a one-hit row is essentially always prose, not headers."""
    flat = [a for al in aliases.values() for a in al]
    for i, row in enumerate(matrix[:window]):
        hits = sum(1 for c in row if c is not None and any(a in _norm(c) for a in flat))
        if hits >= min_hits:
            return i
    return 0

def _build_colmap(headers, aliases=None):
    """Map each canonical field to the best-matching source column index. Aliases are tried
    in DECLARED list order (most specific/correct first) — not sorted by string length. A
    length sort looks appealing ('unit cost' should beat 'cost') but silently breaks whenever
    a longer alias is actually the WRONG match for a shorter, more correct one that's also
    present (e.g. 'fnsku' is longer than 'asin'/'sku' but is never the identifier a tenant's
    known SKUs are keyed by) — see the 'asin' alias list above for the sharpest example."""
    aliases = aliases or _ALIASES
    norm = [_norm(h) for h in headers]
    colmap = {}
    for field, al in aliases.items():
        for alias in al:
            hit = next((i for i, h in enumerate(norm) if alias in h), None)
            if hit is not None and hit not in colmap.values():
                colmap[field] = hit
                break
    return colmap

def read_table(filename, data, aliases=None):
    """Return (headers, body_rows) for a CSV or XLSX upload. Finds the real header row for
    BOTH formats by scanning the first rows for recognizable column names — a plain CSV
    export can carry the same kind of banner/legend preamble an XLSX sheet can. `aliases`
    lets each caller (report_parse.A, cogs._ALIASES, this module's own _ALIASES) drive header
    detection with its own vocabulary; defaults to this module's _ALIASES."""
    aliases = aliases or _ALIASES
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xltx")):
        try:
            import openpyxl
        except ImportError:
            return [], []
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        matrix = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
    else:
        text = data.decode("utf-8-sig", errors="replace") if isinstance(data, (bytes, bytearray)) else data
        matrix = list(csv.reader(io.StringIO(text)))
    if not matrix:
        return [], []
    hdr = _find_header_row(matrix, aliases)
    return matrix[hdr], matrix[hdr + 1:]

def _build_colmap_legacy(headers):
    return _build_colmap(headers)

def _to_float(v):
    if v in (None, ""): return 0.0
    s = re.sub(r"[^0-9.\-]", "", str(v))         # strip currency, commas, spaces
    try: return float(s) if s not in ("", "-", ".") else 0.0
    except ValueError: return 0.0

def _rows_from_matrix(headers, body, channel):
    colmap = _build_colmap(headers)
    if "asin" not in colmap:
        return [], "no ASIN/SKU column found"
    out = []
    for r in body:
        def cell(f):
            i = colmap.get(f)
            return r[i] if (i is not None and i < len(r)) else None
        asin = (str(cell("asin") or "")).strip()
        if not asin or asin.lower() in ("asin", "sku", "nan"):
            continue
        price = _to_float(cell("price"))
        cogs  = _to_float(cell("cogs"))
        if cogs <= 0 and price > 0:
            cogs = round(price * DEFAULT_COGS_RATIO, 2)   # derive, don't drop
        cat = (str(cell("category") or "").strip()) or DEFAULT_CATEGORY
        out.append({
            "asin": asin,
            "title": (str(cell("title") or "")).strip(),
            "price": price,
            "cogs": cogs,
            "category": cat,
            "channel": channel,
        })
    return out, None

def parse_csv_bytes(data, channel="amazon"):
    text = data.decode("utf-8-sig", errors="replace") if isinstance(data, (bytes, bytearray)) else data
    rdr = list(csv.reader(io.StringIO(text)))
    if not rdr: return [], "empty file"
    hdr = _find_header_row(rdr, _ALIASES)
    return _rows_from_matrix(rdr[hdr], rdr[hdr + 1:], channel)

def parse_xlsx_bytes(data, channel="amazon"):
    try:
        import openpyxl
    except ImportError:
        return [], "openpyxl not installed (pip install openpyxl)"
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    matrix = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    if not matrix: return [], "empty sheet"
    hdr_idx = _find_header_row(matrix, _ALIASES)
    return _rows_from_matrix(matrix[hdr_idx], matrix[hdr_idx + 1:], channel)

def parse_upload(filename, data, channel=None):
    """Dispatch by extension. channel defaults to the filename stem."""
    ch = (channel or re.sub(r"\.[^.]+$", "", filename or "store")).strip().lower() or "amazon"
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm", ".xltx")):
        return parse_xlsx_bytes(data, ch)
    return parse_csv_bytes(data, ch)

def parse_many(files):
    """files: list of (filename, bytes, channel|None). Returns (rows, report)."""
    all_rows, report = [], []
    for fn, data, ch in files:
        rows, err = parse_upload(fn, data, ch)
        report.append({"file": fn, "channel": (ch or fn), "rows": len(rows), "error": err})
        all_rows.extend(rows)
    # de-dupe by (asin, channel): last write wins
    seen = {}
    for r in all_rows:
        seen[(r["asin"], r["channel"])] = r
    return list(seen.values()), report
